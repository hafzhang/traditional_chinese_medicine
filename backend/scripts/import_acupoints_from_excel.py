"""
导入穴位数据从 Excel 文件

Excel 文件列映射:
1. 穴位名称 -> name
2. 穴位别名 -> aliases
3. 穴位国际通用代码 -> code
4. 穴位隶属经脉 -> meridian
5. 穴位隶属经脉五行 -> five_element
6. 穴位释义 -> explanation
7. 穴位名称 (重复) -> skip
8. 穴位功能 -> functions
9. 穴位定位和取穴方法 -> location / simple_location
10. 穴位治法-按摩及其它 -> massage_method
11. 穴位治法-灸法 -> moxibustion_method
12. 穴位治法-针刺 -> acupuncture_method
13. 穴位解剖位置和结构 -> anatomical_structure
14. 穴位主要配伍 -> combinations
"""

import sys
import os
import openpyxl
from sqlalchemy.orm import Session

# 添加项目根目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from api.database import engine, SessionLocal, Base
from api.models import Acupoint
from api.data.acupoints import BODY_PART_MERIDIAN_MAP, MERIDIAN_NAME_MAP


def parse_aliases(aliases_str: str) -> list:
    """解析别名字符串为列表"""
    if not aliases_str:
        return []
    # 移除"。"等标点符号，按逗号分隔
    aliases_str = aliases_str.replace("，", ",").replace("。", "").strip()
    if not aliases_str:
        return []
    return [a.strip() for a in aliases_str.split(",") if a.strip()]


def get_pinyin_from_code(code: str) -> str:
    """从代码推断拼音（简化版）"""
    # 这里可以集成拼音库，暂时返回空
    return ""


def get_body_part_from_meridian(meridian: str) -> str:
    """根据经络推断部位"""
    if not meridian:
        return ""

    # 简单映射规则
    if "手太阴肺经" in meridian or "手阳明大肠经" in meridian:
        if "手" in meridian and ("臂" in meridian or "前臂" in meridian):
            return "arm_lower"
        return "arm_upper"
    elif "手少阴心经" in meridian or "手太阳小肠经" in meridian or "手厥阴心包经" in meridian or "手少阳三焦经" in meridian:
        return "arm_upper"
    elif "足阳明胃经" in meridian:
        if "头" in meridian:
            return "head"
        elif "大腿" in meridian or "髀" in meridian:
            return "thigh_upper"
        elif "小腿" in meridian:
            return "thigh_lower"
        elif "足" in meridian:
            return "foot"
        return "abdomen"
    elif "足太阴脾经" in meridian or "足厥阴肝经" in meridian:
        return "thigh_upper"
    elif "足太阳膀胱经" in meridian:
        if "背" in meridian:
            return "back"
        return "thigh_lower"
    elif "足少阳胆经" in meridian:
        if "头" in meridian:
            return "head"
        return "thigh_upper"
    elif "足少阴肾经" in meridian:
        return "thigh_lower"
    elif "督脉" in meridian:
        if "背" in meridian:
            return "back"
        return "head"
    elif "任脉" in meridian:
        if "胸" in meridian:
            return "chest"
        return "abdomen"

    # 默认映射
    for body_part, meridians in BODY_PART_MERIDIAN_MAP.items():
        if meridian in meridians:
            return body_part
    return ""


def import_acupoints_from_excel(file_path: str, db: Session):
    """从 Excel 导入穴位数据"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    imported = 0
    skipped = 0
    errors = []

    # 从第二行开始（第一行是标题）
    for row_idx in range(2, ws.max_row + 1):
        try:
            # 读取各列数据
            name = ws.cell(row_idx, 1).value
            aliases_str = ws.cell(row_idx, 2).value
            code = ws.cell(row_idx, 3).value
            meridian = ws.cell(row_idx, 4).value
            five_element = ws.cell(row_idx, 5).value
            explanation = ws.cell(row_idx, 6).value
            # 第7列是重复的名称，跳过
            functions = ws.cell(row_idx, 8).value
            location_method = ws.cell(row_idx, 9).value
            massage_method = ws.cell(row_idx, 10).value
            moxibustion_method = ws.cell(row_idx, 11).value
            # 针刺方法已移除，跳过第12列
            anatomical_structure = ws.cell(row_idx, 13).value
            combinations = ws.cell(row_idx, 14).value

            if not name:
                skipped += 1
                continue

            # 检查是否已存在
            existing = db.query(Acupoint).filter(Acupoint.name == name).first()
            if existing:
                print(f"  跳过已存在: {name}")
                skipped += 1
                continue

            # 解析别名
            aliases = parse_aliases(aliases_str) if aliases_str else []

            # 处理定位和取穴方法
            location = location_method if location_method else ""

            # 推断部位
            body_part = get_body_part_from_meridian(meridian) if meridian else ""

            # 创建穴位记录
            acupoint = Acupoint(
                name=name.strip(),
                code=code.strip() if code else "",
                pinyin=get_pinyin_from_code(code) if code else "",
                aliases=aliases,
                meridian=meridian.strip() if meridian else "",
                five_element=five_element.strip() if five_element else "",
                body_part=body_part,
                location=location,
                simple_location="",
                explanation=explanation.strip() if explanation else "",
                functions=functions.strip() if functions else "",
                efficacy=[],  # 可以从 functions 中提取
                indications=[],  # 可以从 functions/explanation 中提取
                massage_method=massage_method.strip() if massage_method else "",
                moxibustion_method=moxibustion_method.strip() if moxibustion_method else "",
                anatomical_structure=anatomical_structure.strip() if anatomical_structure else "",
                combinations=combinations.strip() if combinations else "",
                suitable_constitutions=[],
                image_url=f"/static/acupoints/{name.strip()}.jpg"
            )

            db.add(acupoint)
            imported += 1
            print(f"  导入: {name} ({code})")

        except Exception as e:
            error_msg = f"第 {row_idx} 行错误: {str(e)}"
            errors.append(error_msg)
            print(f"  {error_msg}")

    # 提交所有更改
    db.commit()

    # 输出统计
    print("\n" + "=" * 50)
    print(f"导入完成!")
    print(f"  成功导入: {imported} 条")
    print(f"  跳过: {skipped} 条")
    print(f"  错误: {len(errors)} 条")
    if errors:
        print("\n错误详情:")
        for error in errors:
            print(f"  - {error}")

    return imported, skipped, len(errors)


if __name__ == "__main__":
    # 确保表存在
    Base.metadata.create_all(bind=engine)

    # 创建数据库会话
    db = SessionLocal()

    try:
        # Excel 文件路径
        excel_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), "穴位（312）.xlsx")

        if not os.path.exists(excel_file):
            print(f"错误: Excel 文件不存在: {excel_file}")
            sys.exit(1)

        print(f"开始导入穴位数据: {excel_file}")
        print("=" * 50)

        imported, skipped, errors = import_acupoints_from_excel(excel_file, db)

    except Exception as e:
        print(f"导入失败: {str(e)}")
        db.rollback()
        raise
    finally:
        db.close()
